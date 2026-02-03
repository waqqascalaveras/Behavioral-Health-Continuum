"""
Discovery script for summarizing structure of files in etl/downloads.
Outputs a Markdown report to etl/discovery_report.md.
"""
import os
import pandas as pd
import json
import glob
from openpyxl import load_workbook
from pprint import pprint

DOWNLOADS_DIR = os.path.join(os.path.dirname(__file__), 'downloads')
REPORT_PATH = os.path.join(os.path.dirname(__file__), 'discovery_report.md')


def summarize_csv(path):
    try:
        df = pd.read_csv(path, nrows=20)
        summary = {
            'type': 'csv',
            'columns': list(df.columns),
            'dtypes': df.dtypes.apply(str).to_dict(),
            'sample': df.head(5).to_dict(orient='records'),
        }
        return summary
    except Exception as e:
        return {'error': str(e)}

def summarize_excel(path):
    try:
        wb = load_workbook(path, read_only=True)
        sheets = wb.sheetnames
        sheet_summaries = {}
        for sheet in sheets:
            ws = wb[sheet]
            rows = list(ws.iter_rows(values_only=True, max_row=6))
            sheet_summaries[sheet] = {
                'columns': rows[0] if rows else [],
                'sample': rows[1:6] if len(rows) > 1 else [],
            }
        return {
            'type': 'excel',
            'sheets': sheets,
            'sheet_summaries': sheet_summaries,
        }
    except Exception as e:
        return {'error': str(e)}

def summarize_geojson(path):
    try:
        with open(path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        features = data.get('features', [])
        props = features[0]['properties'] if features else {}
        return {
            'type': 'geojson',
            'properties': list(props.keys()),
            'sample': props,
        }
    except Exception as e:
        return {'error': str(e)}

def summarize_rdata(path):
    return {'type': 'rdata', 'note': 'RData summary not implemented. Use R or pyreadr for details.'}

def main():
    files = glob.glob(os.path.join(DOWNLOADS_DIR, '*'))
    report_lines = ['# Discovery Report\n']
    for file in files:
        fname = os.path.basename(file)
        ext = fname.lower().split('.')[-1]
        report_lines.append(f'## {fname}\n')
        if ext == 'csv':
            summary = summarize_csv(file)
        elif ext == 'xlsx':
            summary = summarize_excel(file)
        elif ext == 'geojson':
            summary = summarize_geojson(file)
        elif ext == 'rdata':
            summary = summarize_rdata(file)
        else:
            summary = {'error': 'Unknown or unsupported file type.'}
        report_lines.append('```json')
        report_lines.append(json.dumps(summary, indent=2, default=str))
        report_lines.append('```\n')
    with open(REPORT_PATH, 'w', encoding='utf-8') as f:
        f.write('\n'.join(report_lines))
    print(f'Report written to {REPORT_PATH}')

if __name__ == '__main__':
    main()
